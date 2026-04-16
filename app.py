# -*- coding: utf-8 -*-
"""Brush - 한국 커뮤니티 웹 크롤러 (Streamlit 웹 UI)"""

import io
from datetime import datetime
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

import config
from crawler.clien import ClienCrawler
from crawler.dcinside import DcinsideCrawler
from crawler.fmkorea import FmkoreaCrawler
from crawler.ppomppu import PpomppuCrawler
from crawler.ruliweb import RuliwebCrawler
from crawler.naver_blog import NaverBlogCrawler


SITES = {
    "클리앙": {
        "class": ClienCrawler,
        "example": "https://www.clien.net/service/board/park",
        "global_search": False,
        "needs_url": True,
    },
    "뽐뿌": {
        "class": PpomppuCrawler,
        "example": "https://www.ppomppu.co.kr/zboard/zboard.php?id=freeboard",
        "global_search": False,
        "needs_url": True,
    },
    "루리웹": {
        "class": RuliwebCrawler,
        "example": "https://bbs.ruliweb.com/community/board/300143",
        "global_search": False,
        "needs_url": True,
    },
    "에펨코리아": {
        "class": FmkoreaCrawler,
        "example": "https://www.fmkorea.com/best",
        "global_search": False,
        "needs_url": True,
    },
    "디시인사이드 (통합검색)": {
        "class": DcinsideCrawler,
        "example": "https://gall.dcinside.com/board/lists/?id=programming",
        "global_search": True,
        "needs_url": False,
    },
    "네이버 블로그 (통합검색)": {
        "class": NaverBlogCrawler,
        "example": "https://blog.naver.com/블로그아이디",
        "global_search": True,
        "needs_url": False,
    },
}


st.set_page_config(
    page_title="Brush - 한국 커뮤니티 크롤러",
    page_icon="🖌",
    layout="wide",
)


def build_excel_bytes(posts, site_name):
    """수집 결과를 엑셀 바이트로 변환합니다 (다운로드용)."""
    wb = Workbook()
    ws_posts = wb.active
    ws_posts.title = "게시글"

    headers = ["사이트", "제목", "URL", "작성자", "IP", "게시일", "내용"]
    ws_posts.append(headers)
    _style_header(ws_posts, len(headers))

    for post in posts:
        content = post.get("content", "")
        if len(content) > 32000:
            content = content[:32000] + "... (잘림)"
        ws_posts.append([
            site_name,
            post.get("title", ""),
            post.get("url", ""),
            post.get("author", ""),
            post.get("author_ip", ""),
            post.get("date", ""),
            content,
        ])

    widths = [12, 50, 40, 15, 15, 20, 60]
    for i, w in enumerate(widths, 1):
        ws_posts.column_dimensions[chr(64 + i)].width = w

    ws_comments = wb.create_sheet("댓글")
    c_headers = ["게시글 제목", "게시글 URL", "댓글 작성자", "IP", "댓글일", "내용"]
    ws_comments.append(c_headers)
    _style_header(ws_comments, len(c_headers))

    for post in posts:
        for comment in post.get("comments", []):
            c_content = comment.get("content", "")
            if len(c_content) > 32000:
                c_content = c_content[:32000] + "... (잘림)"
            ws_comments.append([
                post.get("title", ""),
                post.get("url", ""),
                comment.get("author", ""),
                comment.get("author_ip", ""),
                comment.get("date", ""),
                c_content,
            ])

    c_widths = [50, 40, 15, 15, 20, 60]
    for i, w in enumerate(c_widths, 1):
        ws_comments.column_dimensions[chr(64 + i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _style_header(ws, num_cols):
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    align = Alignment(horizontal="center")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = align


# ============================================================
# UI
# ============================================================

st.title("🖌 Brush - 한국 커뮤니티 웹 크롤러")
st.caption("여러 한국 커뮤니티 사이트에서 키워드가 포함된 게시글과 댓글을 수집해서 엑셀로 정리합니다.")

with st.sidebar:
    st.header("⚙️ 설정")

    site_name = st.selectbox("사이트 선택", list(SITES.keys()))
    site = SITES[site_name]

    st.markdown("---")

    keywords_input = st.text_input(
        "키워드 (쉼표로 구분)",
        placeholder="예: 맛집, 추천",
        help="여러 키워드는 쉼표로 구분하세요. 비워두면 전체 게시글을 가져옵니다.",
    )

    if site["global_search"]:
        st.info("통합검색 사이트입니다. 키워드만 입력해도 됩니다.")
        board_url = st.text_input(
            "게시판 URL (선택)",
            placeholder=site["example"],
            help="비워두면 전체 사이트에서 키워드를 검색합니다.",
        )
    else:
        board_url = st.text_input(
            "게시판 URL *",
            placeholder=site["example"],
            help="크롤링할 게시판의 주소를 입력하세요.",
        )

    max_pages = st.slider("최대 페이지 수", 1, 10, 3, help="더 많을수록 시간이 오래 걸립니다.")

    st.markdown("---")
    start = st.button("🚀 크롤링 시작", type="primary", use_container_width=True)


# 메인 영역
if not start:
    st.info("👈 왼쪽에서 사이트와 키워드를 설정한 뒤 **크롤링 시작** 버튼을 눌러주세요.")

    with st.expander("📌 사용 안내"):
        st.markdown("""
        **사용 방법**
        1. 왼쪽 사이드바에서 크롤링할 사이트를 선택합니다.
        2. 키워드를 입력합니다 (여러 개는 쉼표로 구분).
        3. 일반 사이트는 게시판 URL을 입력해야 합니다. 통합검색 사이트는 선택사항입니다.
        4. 최대 페이지 수를 조정하고 **크롤링 시작** 버튼을 누릅니다.
        5. 결과 테이블을 확인하고 엑셀 파일로 다운로드할 수 있습니다.

        **⚠️ 주의사항**
        - 개인 학습 / 연구 목적으로만 사용하세요.
        - 수집한 개인정보(닉네임, IP)는 개인정보보호법의 보호를 받습니다.
        - 요청 간격이 자동으로 조절되므로 너무 많은 페이지를 한번에 요청하지 마세요.
        """)

    with st.expander("✨ 지원 사이트"):
        for name, info in SITES.items():
            tag = "🌐 전체 검색" if info["global_search"] else "📋 게시판 검색"
            st.markdown(f"- **{name}** {tag}")

else:
    # 입력 검증
    if not site["global_search"] and not board_url:
        st.error("❌ 이 사이트는 게시판 URL이 필수입니다.")
        st.stop()

    keywords = [k.strip() for k in keywords_input.split(",") if k.strip()]

    if site["global_search"] and not board_url and not keywords:
        st.error("❌ 통합검색 사이트는 URL이나 키워드 중 하나는 입력해야 합니다.")
        st.stop()

    # 크롤링 실행
    with st.status("🔍 크롤링 중...", expanded=True) as status:
        try:
            crawler = site["class"]()
            st.write(f"**사이트:** {site_name}")
            if board_url:
                st.write(f"**URL:** {board_url}")
            if keywords:
                st.write(f"**키워드:** {', '.join(keywords)}")
            st.write(f"**최대 페이지:** {max_pages}")

            # 크롤링 수행
            with st.spinner("게시글을 수집하는 중입니다. 잠시만 기다려주세요..."):
                results = crawler.crawl(
                    board_url=board_url or "",
                    keywords=keywords,
                    max_pages=max_pages,
                )

            total_comments = sum(len(p.get("comments", [])) for p in results)
            status.update(
                label=f"✅ 완료! 게시글 {len(results)}개, 댓글 {total_comments}개 수집",
                state="complete",
                expanded=False,
            )

        except Exception as e:
            status.update(label=f"❌ 오류 발생: {e}", state="error")
            st.exception(e)
            st.stop()

    if not results:
        st.warning("🔍 수집된 게시글이 없습니다. 키워드나 URL을 확인해주세요.")
        st.stop()

    # 결과 요약
    col1, col2, col3 = st.columns(3)
    col1.metric("수집 게시글", f"{len(results)}개")
    col2.metric("수집 댓글", f"{total_comments}개")
    col3.metric("대상 사이트", site_name)

    # 다운로드 버튼
    st.markdown("### 📥 다운로드")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_bytes = build_excel_bytes(results, site_name)
    st.download_button(
        label="📊 엑셀 파일로 다운로드",
        data=excel_bytes,
        file_name=f"crawl_{site_name}_{timestamp}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    # 게시글 표
    st.markdown("### 📝 게시글")
    posts_data = []
    for p in results:
        posts_data.append({
            "제목": p.get("title", ""),
            "작성자": p.get("author", ""),
            "IP": p.get("author_ip", ""),
            "게시일": p.get("date", ""),
            "댓글수": len(p.get("comments", [])),
            "URL": p.get("url", ""),
        })
    df_posts = pd.DataFrame(posts_data)
    st.dataframe(df_posts, use_container_width=True, hide_index=True)

    # 게시글별 상세 내용
    st.markdown("### 🔎 게시글 상세")
    for i, p in enumerate(results):
        with st.expander(f"**{i + 1}. {p.get('title', '')[:80]}**"):
            info_cols = st.columns(3)
            info_cols[0].markdown(f"**작성자:** {p.get('author', '')}")
            info_cols[1].markdown(f"**IP:** {p.get('author_ip', '-')}")
            info_cols[2].markdown(f"**날짜:** {p.get('date', '')}")
            st.markdown(f"[원문 링크]({p.get('url', '')})")

            st.markdown("**본문:**")
            content = p.get("content", "")
            if len(content) > 2000:
                st.text(content[:2000] + "...")
            else:
                st.text(content)

            comments = p.get("comments", [])
            if comments:
                st.markdown(f"**댓글 ({len(comments)}개):**")
                comment_data = [{
                    "작성자": c.get("author", ""),
                    "IP": c.get("author_ip", ""),
                    "날짜": c.get("date", ""),
                    "내용": c.get("content", ""),
                } for c in comments]
                st.dataframe(pd.DataFrame(comment_data), use_container_width=True, hide_index=True)
