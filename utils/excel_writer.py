import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import config


def save_to_excel(posts, site_name):
    """크롤링 결과를 엑셀 파일로 저장합니다.

    시트1 '게시글': 사이트명, 제목, URL, 작성자, IP, 게시일, 내용
    시트2 '댓글': 게시글 제목, URL, 댓글 작성자, IP, 댓글일, 내용

    반환: 저장된 파일 경로
    """
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"crawl_results_{site_name}_{timestamp}.xlsx"
    filepath = os.path.join(config.OUTPUT_DIR, filename)

    wb = Workbook()

    # --- 시트1: 게시글 ---
    ws_posts = wb.active
    ws_posts.title = "게시글"

    post_headers = ["사이트", "제목", "URL", "작성자", "IP", "게시일", "내용"]
    ws_posts.append(post_headers)
    _style_header_row(ws_posts, len(post_headers))

    for post in posts:
        content = post.get("content", "")
        # 엑셀 셀 최대 길이 제한 (32767자)
        if len(content) > 32000:
            content = content[:32000] + "... (내용이 너무 길어 잘림)"

        ws_posts.append([
            site_name,
            post.get("title", ""),
            post.get("url", ""),
            post.get("author", ""),
            post.get("author_ip", ""),
            post.get("date", ""),
            content,
        ])

    # 열 너비 조정
    ws_posts.column_dimensions["A"].width = 12
    ws_posts.column_dimensions["B"].width = 50
    ws_posts.column_dimensions["C"].width = 40
    ws_posts.column_dimensions["D"].width = 15
    ws_posts.column_dimensions["E"].width = 15
    ws_posts.column_dimensions["F"].width = 20
    ws_posts.column_dimensions["G"].width = 60

    # --- 시트2: 댓글 ---
    ws_comments = wb.create_sheet("댓글")

    comment_headers = ["게시글 제목", "게시글 URL", "댓글 작성자", "IP", "댓글일", "내용"]
    ws_comments.append(comment_headers)
    _style_header_row(ws_comments, len(comment_headers))

    for post in posts:
        for comment in post.get("comments", []):
            comment_content = comment.get("content", "")
            if len(comment_content) > 32000:
                comment_content = comment_content[:32000] + "... (내용이 너무 길어 잘림)"

            ws_comments.append([
                post.get("title", ""),
                post.get("url", ""),
                comment.get("author", ""),
                comment.get("author_ip", ""),
                comment.get("date", ""),
                comment_content,
            ])

    ws_comments.column_dimensions["A"].width = 50
    ws_comments.column_dimensions["B"].width = 40
    ws_comments.column_dimensions["C"].width = 15
    ws_comments.column_dimensions["D"].width = 15
    ws_comments.column_dimensions["E"].width = 20
    ws_comments.column_dimensions["F"].width = 60

    wb.save(filepath)
    return filepath


def _style_header_row(ws, num_cols):
    """헤더 행에 스타일을 적용합니다."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
